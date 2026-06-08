import os
import calendar as _calendar
import datetime as dt
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.template import loader
from django.urls import reverse_lazy
from django.views.generic.edit import DeleteView
from django.db.models import Q, Count, OuterRef, Exists
from datetime import datetime

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.auth.models import User

from apps.superadmin.decorators import edu_admin_required, monitoring_required
from apps.superadmin.models import Administrator
from apps.main.models import Location
from .models import Group, Direction, Student, Smena, SmenaSlot, GroupLesson, GroupSchedule, StudentAttendance, MONTH_CHOICES
from .forms import GroupForm, DirectionForm, SmenaForm, SmenaSlotFormSet


def _generate_password():
    """4 ta belgi: faqat unli harflar va raqamlar"""
    chars = 'aeiou0123456789'
    return ''.join(random.choices(chars, k=4))


def _make_login(student_id):
    """8 xonali login: student PK ni 0 bilan to'ldirish"""
    return str(student_id).zfill(8)


MONTH_NAMES = {
    1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
    5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
    9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr'
}


def _get_admin_filial(request):
    """Admin va uning filialini qaytaradi"""
    admin_user = Administrator.objects.get(user=request.user)
    filial_id = request.session.get('selected_filial_id')
    filial_id = filial_id if filial_id and filial_id != 'super_admin' else None
    if not filial_id and admin_user.filial_id:
        filial_id = admin_user.filial_id
    return admin_user, filial_id


# ============================================================
# GURUHLAR
# ============================================================

@edu_admin_required
def groups_list(request):
    admin_user, filial_id = _get_admin_filial(request)
    now = datetime.now()

    # Filter parametrlari — default joriy yil va oy
    filter_year      = request.GET.get('year',      now.year)
    filter_month     = request.GET.get('month',     now.month)
    filter_direction = request.GET.get('direction', '')
    search_query     = request.GET.get('q', '')

    try:
        filter_year  = int(filter_year)
        filter_month = int(filter_month)
    except (ValueError, TypeError):
        filter_year  = now.year
        filter_month = now.month

    # Lokatsiya (faol jadval) bor-yo'qligini aniqlash uchun subquery
    has_schedule_sq = GroupSchedule.objects.filter(
        group=OuterRef('pk'),
        is_active=True,
    )

    groups = Group.objects.filter(
        organization=admin_user.organization,
        year=filter_year,
        month=filter_month,
    ).select_related('filial', 'direction').annotate(
        total_students=Count('students', distinct=True),
        registered_students=Count(
            'students',
            filter=Q(students__is_registered=True),
            distinct=True,
        ),
        has_schedule=Exists(has_schedule_sq),
    ).order_by('name')

    if filial_id:
        groups = groups.filter(filial_id=filial_id)

    if filter_direction:
        groups = groups.filter(direction_id=filter_direction)

    if search_query:
        groups = groups.filter(Q(name__icontains=search_query))

    # Yo'nalishlar filter uchun (faqat admin filiali)
    directions = Direction.objects.filter(organization=admin_user.organization)
    if filial_id:
        directions = directions.filter(filial_id=filial_id)

    # Yillar ro'yxati (tanlash uchun) — mavjud guruhlardan
    years = sorted(
        Group.objects.filter(organization=admin_user.organization)
        .values_list('year', flat=True).distinct(),
        reverse=True
    )
    if now.year not in years:
        years.insert(0, now.year)

    paginator = Paginator(groups, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj':        page_obj,
        'segment':         'groups',
        'search_query':    search_query,
        'filter_year':     filter_year,
        'filter_month':    filter_month,
        'filter_direction': filter_direction,
        'month_name':      MONTH_NAMES[filter_month],
        'month_choices':   MONTH_CHOICES,
        'years':           years,
        'directions':      directions,
    }
    return HttpResponse(loader.get_template('home/students/groups_list.html').render(context, request))


@edu_admin_required
def group_create(request):
    admin_user, filial_id = _get_admin_filial(request)

    if request.method == 'POST':
        form = GroupForm(request.POST, filial_id=filial_id)
        if form.is_valid():
            group = form.save(commit=False)
            group.organization = admin_user.organization
            group.filial_id = filial_id
            group.save()
            return redirect('groups_list')
    else:
        form = GroupForm(filial_id=filial_id)

    context = {
        'form':    form,
        'segment': 'groups',
    }
    return render(request, 'home/students/group_create.html', context)


@edu_admin_required
def group_detail(request, pk):
    group      = get_object_or_404(Group, pk=pk)
    admin_user, filial_id = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group, filial_id=group.filial_id)
        if form.is_valid():
            form.save()
            return redirect('groups_list')
    else:
        form = GroupForm(instance=group, filial_id=group.filial_id)

    context = {
        'form':    form,
        'group':   group,
        'segment': 'groups',
    }
    return render(request, 'home/students/group_detail.html', context)


@edu_admin_required
def directions_by_filial(request):
    """AJAX: filialni o'zgartirganda yo'nalishlar ro'yxatini qaytaradi"""
    filial_id = request.GET.get('filial_id')
    directions = Direction.objects.filter(filial_id=filial_id).values('id', 'name')
    return JsonResponse({'directions': list(directions)})


class GroupDelete(DeleteView):
    model = Group
    success_url = reverse_lazy('groups_list')
    template_name = 'home/students/group_confirm_delete.html'


# ============================================================
# YO'NALISHLAR
# ============================================================

@edu_admin_required
def directions_list(request):
    admin_user, filial_id = _get_admin_filial(request)

    directions = Direction.objects.filter(organization=admin_user.organization).order_by('name')
    if filial_id:
        directions = directions.filter(filial_id=filial_id)

    search_query = request.GET.get('q', '')
    if search_query:
        directions = directions.filter(Q(name__icontains=search_query))

    paginator  = Paginator(directions, 20)
    page_obj   = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj':     page_obj,
        'segment':      'directions',
        'search_query': search_query,
    }
    return HttpResponse(loader.get_template('home/students/directions_list.html').render(context, request))


@edu_admin_required
def direction_create(request):
    admin_user, filial_id = _get_admin_filial(request)

    if request.method == 'POST':
        form = DirectionForm(request.POST)
        if form.is_valid():
            direction = form.save(commit=False)
            direction.organization = admin_user.organization
            direction.filial_id = filial_id
            direction.save()
            return redirect('directions_list')
    else:
        form = DirectionForm()

    context = {
        'form':    form,
        'segment': 'directions',
    }
    return render(request, 'home/students/direction_create.html', context)


@edu_admin_required
def direction_detail(request, pk):
    direction  = get_object_or_404(Direction, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if direction.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        form = DirectionForm(request.POST, instance=direction)
        if form.is_valid():
            form.save()
            return redirect('directions_list')
    else:
        form = DirectionForm(instance=direction)

    context = {
        'form':      form,
        'direction': direction,
        'segment':   'directions',
    }
    return render(request, 'home/students/direction_detail.html', context)


# ============================================================
# TAKLIF HAVOLALARI
# ============================================================

@edu_admin_required
def invite_links(request):
    admin_user, filial_id = _get_admin_filial(request)
    now = datetime.now()

    filter_year      = request.GET.get('year',  now.year)
    filter_month     = request.GET.get('month', now.month)
    filter_direction = request.GET.get('direction', '')

    try:
        filter_year  = int(filter_year)
        filter_month = int(filter_month)
    except (ValueError, TypeError):
        filter_year  = now.year
        filter_month = now.month

    groups = Group.objects.filter(
        organization=admin_user.organization,
        year=filter_year,
        month=filter_month,
    ).select_related('filial', 'direction')

    if filial_id:
        groups = groups.filter(filial_id=filial_id)

    if filter_direction:
        groups = groups.filter(direction_id=filter_direction)

    directions = Direction.objects.filter(organization=admin_user.organization)
    if filial_id:
        directions = directions.filter(filial_id=filial_id)

    years = sorted(
        Group.objects.filter(organization=admin_user.organization)
        .values_list('year', flat=True).distinct(),
        reverse=True
    )
    if now.year not in years:
        years.insert(0, now.year)

    bot_username = os.environ.get('BOT_USERNAME', '')

    groups_with_links = [
        (g, f"https://t.me/{bot_username}?start=grp_{g.invite_token}")
        for g in groups
    ]

    context = {
        'groups_with_links': groups_with_links,
        'segment':           'invites',
        'filter_year':       filter_year,
        'filter_month':      filter_month,
        'filter_direction':  filter_direction,
        'month_choices':     MONTH_CHOICES,
        'years':             years,
        'directions':        directions,
    }
    return render(request, 'home/students/invite_links.html', context)


@edu_admin_required
def regenerate_invite_token(request, pk):
    if request.method != 'POST':
        return redirect('invite_links')
    group      = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)
    import uuid as _uuid
    group.invite_token = _uuid.uuid4()
    group.save(update_fields=['invite_token'])
    return redirect(request.META.get('HTTP_REFERER', 'invite_links'))


class DirectionDelete(DeleteView):
    model = Direction
    success_url = reverse_lazy('directions_list')
    template_name = 'home/students/direction_confirm_delete.html'


# ============================================================
# GURUH TINGLOVCHILARI
# ============================================================

@edu_admin_required
def group_students(request, pk):
    group      = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    error   = None
    success = None

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            error = "Fayl tanlanmagan."
        elif not excel_file.name.endswith(('.xlsx', '.xls')):
            error = "Faqat .xlsx yoki .xls formatidagi fayl qabul qilinadi."
        else:
            try:
                wb   = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                ws   = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))

                added = 0
                skipped = 0
                for row in rows:
                    if not row or not row[0]:
                        continue
                    full_name = str(row[0]).strip()
                    phone     = str(row[1]).strip() if len(row) > 1 and row[1] else None

                    student, created = Student.objects.get_or_create(
                        full_name=full_name,
                        organization=admin_user.organization,
                        defaults={
                            'filial': group.filial,
                            'phone':  phone,
                        }
                    )

                    # Yangi tinglovchi uchun user yaratish
                    if created or not student.user_id:
                        login    = _make_login(student.pk)
                        password = _generate_password()
                        user     = User.objects.create_user(username=login, password=password)
                        student.user           = user
                        student.plain_password = password
                        student.save(update_fields=['user', 'plain_password'])

                    if group.students.filter(pk=student.pk).exists():
                        skipped += 1
                    else:
                        group.students.add(student)
                        added += 1

                success = f"{added} ta tinglovchi qo'shildi. {skipped} ta allaqachon guruhda."
            except Exception as e:
                error = f"Faylni o'qishda xatolik: {e}"

    students   = group.students.all().order_by('full_name')
    paginator  = Paginator(students, 30)
    page_obj   = paginator.get_page(request.GET.get('page'))

    context = {
        'group':   group,
        'page_obj': page_obj,
        'segment': 'groups',
        'error':   error,
        'success': success,
    }
    return render(request, 'home/students/group_students.html', context)


@edu_admin_required
def group_students_export(request, pk):
    group      = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tinglovchilar"

    # Styles
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='CCCCCC')
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'F.I.Sh', 'Telefon', 'Login (ID)', 'Parol']
    widths  = [5, 35, 18, 14, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = brd
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    students = group.students.all().order_by('full_name')
    for i, st in enumerate(students, 1):
        login = _make_login(st.pk) if st.pk else ''
        row_data = [i, st.full_name, st.phone or '', login, st.plain_password or '']
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = brd
            cell.alignment = Alignment(vertical='center')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"group_{group.pk}_students.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@edu_admin_required
def mark_muqobil(request, pk):
    """
    POST: Guruh tinglovchilaridan bir nechtasini muqobil malaka oshirishga o'tkazish.
    student_ids — checkbox orqali tanlangan IDlar ro'yxati
    muqobil_from — sana (YYYY-MM-DD)
    muqobil_to — avtomatik: guruh oyi oxirgi kuni
    """
    import calendar as _cal
    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        muqobil_from_str = request.POST.get('muqobil_from', '')

        error = None
        try:
            muqobil_date = dt.date.fromisoformat(muqobil_from_str) if muqobil_from_str else dt.date.today()
        except ValueError:
            error = "Noto'g'ri sana formati."
            muqobil_date = dt.date.today()

        if not error and student_ids:
            # Muqobil tugash sanasi: guruh oyi + yilining oxirgi kuni
            last_day = _cal.monthrange(group.year, group.month)[1]
            muqobil_to_date = dt.date(group.year, group.month, last_day)

            Student.objects.filter(
                id__in=student_ids,
                groups=group,
            ).update(
                is_muqobil=True,
                muqobil_from=muqobil_date,
                muqobil_to=muqobil_to_date,
            )

    return redirect('group_students', pk=pk)


@edu_admin_required
def mark_masofaviy(request, pk):
    """
    POST: Guruh tinglovchilaridan bir nechtasini masofaviy ta'limga o'tkazish.
    student_ids  — checkbox orqali tanlangan IDlar
    masofaviy_from — boshlanish sanasi (YYYY-MM-DD)
    masofaviy_to   — tugash sanasi (YYYY-MM-DD), ixtiyoriy
    """
    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        student_ids       = request.POST.getlist('student_ids')
        from_str          = request.POST.get('masofaviy_from', '')
        to_str            = request.POST.get('masofaviy_to', '')

        try:
            from_date = dt.date.fromisoformat(from_str) if from_str else dt.date.today()
        except ValueError:
            from_date = dt.date.today()

        try:
            to_date = dt.date.fromisoformat(to_str) if to_str else None
        except ValueError:
            to_date = None

        if student_ids:
            Student.objects.filter(
                id__in=student_ids,
                groups=group,
            ).update(
                is_masofaviy=True,
                masofaviy_from=from_date,
                masofaviy_to=to_date,
            )

    return redirect('group_students', pk=pk)


@edu_admin_required
def group_student_remove(request, pk, student_pk):
    group      = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        group.students.remove(student_pk)

    return redirect('group_students', pk=pk)


@edu_admin_required
def student_clear(request, pk, student_pk):
    """
    Tinglovchining barcha ma'lumotlarini tozalash —
    noto'g'ri tanlangan bo'lsa, qaytadan boshlash uchun.
    Tozalanadi: telegram_id, face_image, face_encoding,
                face_verified, is_registered, registration_status,
                barcha StudentAttendance yozuvlari.
    """
    import os
    group      = get_object_or_404(Group, pk=pk)
    student    = get_object_or_404(Student, pk=student_pk)
    admin_user, _ = _get_admin_filial(request)

    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    if request.method == 'POST':
        # Yuz rasmini diskdan o'chirish
        if student.face_image:
            try:
                if os.path.isfile(student.face_image.path):
                    os.remove(student.face_image.path)
            except Exception:
                pass

        # Barcha davomat yozuvlarini o'chirish
        StudentAttendance.objects.filter(student=student).delete()

        # Student maydonlarini tozalash
        student.telegram_id         = None
        student.face_image          = None
        student.face_encoding       = None
        student.face_verified       = False
        student.is_registered       = False
        student.registration_status = 'pending'
        student.save(update_fields=[
            'telegram_id', 'face_image', 'face_encoding',
            'face_verified', 'is_registered', 'registration_status',
        ])

    return redirect('group_students', pk=pk)


# ============================================================
# SMENALAR
# ============================================================

@edu_admin_required
def smenas_list(request):
    admin_user, filial_id = _get_admin_filial(request)
    smenas = Smena.objects.filter(organization=admin_user.organization)
    if filial_id:
        smenas = smenas.filter(filial_id=filial_id)
    return render(request, 'home/students/smenas_list.html', {
        'smenas': smenas,
        'segment': 'smenas',
    })


@edu_admin_required
def smena_create(request):
    admin_user, filial_id = _get_admin_filial(request)
    if request.method == 'POST':
        form = SmenaForm(request.POST)
        formset = SmenaSlotFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            smena = form.save(commit=False)
            smena.organization = admin_user.organization
            smena.filial_id = filial_id
            smena.save()
            formset.instance = smena
            formset.save()
            return redirect('smenas_list')
    else:
        form = SmenaForm()
        formset = SmenaSlotFormSet()
    return render(request, 'home/students/smena_form.html', {
        'form': form, 'formset': formset, 'segment': 'smenas'
    })


@edu_admin_required
def smena_detail(request, pk):
    smena = get_object_or_404(Smena, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if smena.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)
    if request.method == 'POST':
        form = SmenaForm(request.POST, instance=smena)
        formset = SmenaSlotFormSet(request.POST, instance=smena)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('smenas_list')
    else:
        form = SmenaForm(instance=smena)
        formset = SmenaSlotFormSet(instance=smena)
    return render(request, 'home/students/smena_form.html', {
        'form': form, 'formset': formset, 'smena': smena, 'segment': 'smenas'
    })


class SmenaDelete(DeleteView):
    model = Smena
    success_url = reverse_lazy('smenas_list')
    template_name = 'home/students/smena_confirm_delete.html'


# ============================================================
# GURUH JADVALI — KALENDAR
# ============================================================

@edu_admin_required
def group_schedule(request, pk):
    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if group.organization != admin_user.organization:
        return HttpResponse("Ruxsatnoma yo'q", status=403)

    # Calendar weeks for the group's month
    cal_weeks = _calendar.monthcalendar(group.year, group.month)

    # Existing lessons for this group
    lessons_qs = GroupLesson.objects.filter(group=group).select_related('location', 'smena')
    lesson_map = {lesson.date: lesson for lesson in lessons_qs}

    limit_start = group.limit_start_date

    # Build calendar data structure
    weeks = []
    for week in cal_weeks:
        week_days = []
        for day in week:
            if day == 0:
                week_days.append(None)
            else:
                date = dt.date(group.year, group.month, day)
                is_before_limit = limit_start is not None and date < limit_start
                is_limit_date   = limit_start is not None and date == limit_start
                week_days.append({
                    'day': day,
                    'date': date.isoformat(),
                    'lesson': lesson_map.get(date),
                    'is_past': date < dt.date.today(),
                    'is_before_limit': is_before_limit,
                    'is_limit_date':   is_limit_date,
                })
        weeks.append(week_days)

    smenas = Smena.objects.filter(organization=admin_user.organization)
    locations = Location.objects.filter(organization=admin_user.organization)

    return render(request, 'home/students/group_schedule.html', {
        'group': group,
        'weeks': weeks,
        'smenas': smenas,
        'locations': locations,
        'segment': 'groups',
        'weekday_names': ['Du', 'Se', 'Cho', 'Pa', 'Ju', 'Sha', 'Ya'],
        'limit_start_date':     limit_start,
        'limit_start_date_str': limit_start.isoformat() if limit_start else '',
    })


@edu_admin_required
def save_group_lessons(request, pk):
    """AJAX: bir yoki bir nechta sanaga smena+lokatsiya biriktirish"""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if group.organization != admin_user.organization:
        return JsonResponse({'ok': False, 'error': 'Ruxsat yo\'q'}, status=403)

    import json
    data = json.loads(request.body)
    dates = data.get('dates', [])      # list of 'YYYY-MM-DD'
    smena_id = data.get('smena_id')
    location_id = data.get('location_id')

    if not dates:
        return JsonResponse({'ok': False, 'error': 'Sana tanlanmagan'})

    smena = Smena.objects.filter(pk=smena_id, organization=admin_user.organization).first() if smena_id else None
    location = Location.objects.filter(pk=location_id, organization=admin_user.organization).first() if location_id else None

    saved = []
    for date_str in dates:
        try:
            date = dt.date.fromisoformat(date_str)
        except ValueError:
            continue
        GroupLesson.objects.update_or_create(
            group=group,
            date=date,
            defaults={'smena': smena, 'location': location},
        )
        saved.append(date_str)

    return JsonResponse({'ok': True, 'saved': saved})


@edu_admin_required
def delete_group_lesson(request, pk, date_str):
    """DELETE: muayyan sananing darsini o'chirish"""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if group.organization != admin_user.organization:
        return JsonResponse({'ok': False}, status=403)

    try:
        date = dt.date.fromisoformat(date_str)
    except ValueError:
        return JsonResponse({'ok': False})

    GroupLesson.objects.filter(group=group, date=date).delete()
    return JsonResponse({'ok': True})


@edu_admin_required
def set_group_limit_date(request, pk):
    """AJAX: Guruh uchun limit hisoblash boshlanish sanasini belgilash yoki tozalash."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    group = get_object_or_404(Group, pk=pk)
    admin_user, _ = _get_admin_filial(request)
    if group.organization != admin_user.organization:
        return JsonResponse({'ok': False, 'error': "Ruxsat yo'q"}, status=403)

    import json as _json
    data = _json.loads(request.body)
    date_str = data.get('date')  # None → tozalash

    if date_str:
        try:
            date = dt.date.fromisoformat(date_str)
        except ValueError:
            return JsonResponse({'ok': False, 'error': "Noto'g'ri sana formati"})
        group.limit_start_date = date
    else:
        group.limit_start_date = None

    group.save(update_fields=['limit_start_date'])
    return JsonResponse({'ok': True, 'date': date_str or None})


# ============================================================
# TINGLOVCHI TELEGRAM ID TOZALASH
# ============================================================

@edu_admin_required
def student_telegram_reset(request):
    """
    Tinglovchini ID yoki ismi bo'yicha qidirish va
    Telegram bog'lanishini tozalash (qayta ro'yxatdan o'tish uchun).
    """
    admin_user, filial_id = _get_admin_filial(request)

    query    = request.GET.get('q', '').strip()
    students = []
    searched = False

    if query:
        searched = True
        qs = Student.objects.filter(organization=admin_user.organization)

        # Raqam bo'lsa — ID bo'yicha ham qidirish
        if query.isdigit():
            qs = qs.filter(
                Q(full_name__icontains=query) | Q(id=int(query))
            )
        else:
            qs = qs.filter(full_name__icontains=query)

        students = qs.order_by('full_name')[:50]

    return render(request, 'home/students/telegram_reset.html', {
        'segment':  'telegram_reset',
        'query':    query,
        'searched': searched,
        'students': students,
    })


@edu_admin_required
def student_telegram_reset_confirm(request, pk):
    """POST: Tinglovchining Telegram ID va user bog'lanishini tozalaydi."""
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('student_telegram_reset')

    admin_user, _ = _get_admin_filial(request)
    student = get_object_or_404(Student, pk=pk, organization=admin_user.organization)

    student.telegram_id = None
    student.user        = None
    student.save(update_fields=['telegram_id', 'user'])

    url = reverse('student_telegram_reset')
    return redirect(f"{url}?q={student.full_name}&reset_done={student.pk}")


def student_web_app(request):
    """Tinglovchi davomat web sahifasi (Telegram WebApp orqali ochiladi)"""
    html_template = loader.get_template('students/web_app_page.html')
    response = HttpResponse(html_template.render({}, request))
    response['X-Frame-Options'] = 'ALLOWALL'
    return response


def edu_admin_web_app(request):
    """O'quv admin davomat web sahifasi (Telegram WebApp orqali ochiladi)"""
    html_template = loader.get_template('students/edu_admin_web_app.html')
    response = HttpResponse(html_template.render({}, request))
    response['X-Frame-Options'] = 'ALLOWALL'
    return response


# ============================================================
# TINGLOVCHILAR HISOBOTI
# ============================================================

def _build_student_report(group, date_from, date_to):
    """
    Guruh tinglovchilari bo'yicha davomat hisobotini qaytaradi.

    Hisob PARA asosida (kun emas):
    - check_in <= para_start          → para keldi (o'z vaqtida)
    - para_start < check_in <= +40min → para keldi (kechikib, kechikdi +1)
    - check_in > para_start + 40min   → paraga kelmadi
    - Davomat yozilmagan kun          → barcha paralarga kelmadi
    """
    from .models import StudentAttendance, GroupLesson
    import datetime as _dt

    today = _dt.date.today()
    effective_to = min(date_to, today)
    LATE_THRESHOLD    = 40   # daqiqa: shu vaqtdan kech kelsa paraga kelmadi hisoblanadi
    PARA_DURATION_MIN = 80   # slot.end yo'q bo'lsa default para davomiyligi (daqiqa)

    lessons_qs = GroupLesson.objects.filter(
        group=group,
        date__gte=date_from,
        date__lte=effective_to,
        smena__isnull=False,
    ).select_related('smena').order_by('date')

    lesson_map = {lesson.date: lesson for lesson in lessons_qs}

    # Jami paralar soni (barcha dars kunlarining paralari yig'indisi)
    total_paras = sum(
        len(l.smena.get_slots()) for l in lesson_map.values()
    )

    students = group.students.order_by('full_name')
    rows = []
    for student in students:
        # Muqobil sanasidan keyingi kunlar bu tinglovchi uchun hisoblanmaydi
        muqobil_date = student.muqobil_from if student.is_muqobil and student.muqobil_from else None

        # Ushbu tinglovchi uchun hisoblash kerak bo'lgan darslar (muqobil sanasigacha)
        if muqobil_date:
            student_lesson_map = {d: l for d, l in lesson_map.items() if d < muqobil_date}
        else:
            student_lesson_map = lesson_map

        student_total_paras = sum(len(l.smena.get_slots()) for l in student_lesson_map.values())

        atts = StudentAttendance.objects.filter(
            student=student,
            group=group,
            date__gte=date_from,
            date__lte=effective_to,
        )
        att_by_date = {att.date: att for att in atts}

        present_count    = 0
        absent_count     = 0
        late_count       = 0
        late_mins_total  = 0
        early_count      = 0
        early_mins_total = 0

        for date, lesson in student_lesson_map.items():
            smena  = lesson.smena
            slots  = smena.get_slots()
            para_count = len(slots)

            att = att_by_date.get(date)

            # Kelmagan yoki davomat yozilmagan — barcha paralar missed
            if att is None or not att.check_in or att.status in ('absent', 'excused'):
                absent_count += para_count
                continue

            check_in_dt  = _dt.datetime.combine(date, att.check_in)
            checkout_dt  = _dt.datetime.combine(date, att.check_out) if att.check_out else None

            for slot in slots:
                para_start_dt = _dt.datetime.combine(date, slot.start)
                para_end_dt   = (
                    _dt.datetime.combine(date, slot.end) if slot.end
                    else para_start_dt + _dt.timedelta(minutes=PARA_DURATION_MIN)
                )

                # Check_out bu paradan oldin bo'lsa → paraga kelmadi
                if checkout_dt and checkout_dt <= para_start_dt:
                    absent_count += 1
                    continue

                # Para tugagandan keyin keldi yoki 40+ daqiqa kech → yo'q
                if check_in_dt >= para_end_dt or check_in_dt > para_start_dt + _dt.timedelta(minutes=LATE_THRESHOLD):
                    absent_count += 1
                    continue

                # Paraga keldi
                present_count += 1

                # Kechikish
                if check_in_dt > para_start_dt:
                    late_count      += 1
                    late_mins_total += int((check_in_dt - para_start_dt).total_seconds() / 60)

                # Erta ketish: check_out para tugashidan oldin
                if checkout_dt and checkout_dt < para_end_dt:
                    early_count      += 1
                    early_mins_total += int((para_end_dt - checkout_dt).total_seconds() / 60)

        percent = round(present_count / student_total_paras * 100) if student_total_paras else 0

        rows.append({
            'student':      student,
            'is_muqobil':   student.is_muqobil,
            'muqobil_from': muqobil_date,
            'present':      present_count,
            'absent':       absent_count,
            'late_count':   late_count,
            'late_mins':    late_mins_total,
            'early_count':  early_count,
            'early_mins':   early_mins_total,
            'percent':      percent,
            'total':        student_total_paras,
        })

    return rows, total_paras


@monitoring_required
def student_report(request):
    admin_user, filial_id = _get_admin_filial(request)
    now = datetime.now()

    # Guruhlar ro'yxati (filter uchun)
    groups_qs = Group.objects.filter(
        organization=admin_user.organization
    ).select_related('filial', 'direction').order_by('-year', '-month', 'name')
    if filial_id:
        groups_qs = groups_qs.filter(filial_id=filial_id)

    group_id = request.GET.get('group_id')
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to',   '')

    selected_group = None
    rows = []
    total_lessons = 0
    error = None

    if group_id:
        selected_group = groups_qs.filter(pk=group_id).first()
        if selected_group:
            # Sana oraligi — default: guruh oyining birinchi va oxirgi kuni
            try:
                date_from = dt.date.fromisoformat(date_from_str) if date_from_str else \
                            dt.date(selected_group.year, selected_group.month, 1)
                last_day  = _calendar.monthrange(selected_group.year, selected_group.month)[1]
                date_to   = dt.date.fromisoformat(date_to_str) if date_to_str else \
                            dt.date(selected_group.year, selected_group.month, last_day)
            except ValueError:
                error = "Sana formati noto'g'ri."
                date_from = dt.date(selected_group.year, selected_group.month, 1)
                last_day  = _calendar.monthrange(selected_group.year, selected_group.month)[1]
                date_to   = dt.date(selected_group.year, selected_group.month, last_day)

            if request.GET.get('export') == 'xlsx':
                return _export_student_report_xlsx(selected_group, date_from, date_to)

            rows, total_lessons = _build_student_report(selected_group, date_from, date_to)
        else:
            error = "Guruh topilmadi."
            date_from = dt.date(now.year, now.month, 1)
            date_to   = dt.date.today()
    else:
        date_from = dt.date(now.year, now.month, 1)
        date_to   = dt.date.today()

    context = {
        'groups':         groups_qs,
        'selected_group': selected_group,
        'rows':           rows,
        'total_lessons':  total_lessons,
        'date_from':      date_from,
        'date_to':        date_to,
        'date_from_str':  date_from.isoformat() if date_from else '',
        'date_to_str':    date_to.isoformat() if date_to else '',
        'error':          error,
        'segment':        'student_report',
    }
    return render(request, 'home/students/student_report.html', context)


def _export_student_report_xlsx(group, date_from, date_to):
    rows, total_lessons = _build_student_report(group, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tinglovchilar hisoboti"

    # Styles
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CCCCCC')
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')

    # Sarlavha
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"{group.name} — Davomat hisoboti  ({date_from} – {date_to})"
    title_cell.font  = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = [
        '#', 'F.I.Sh',
        f'Jami para\n({total_lessons})', 'Keldi\n(para)', 'Kelmadi\n(para)',
        'Kechikdi (marta)', 'Kechikish (daqiqa)',
        'Erta ketdi (marta)', 'Erta ketish (daqiqa)',
        'Davomat %'
    ]
    widths = [5, 32, 12, 10, 10, 18, 18, 18, 18, 12]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = brd
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 32

    muqobil_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    for i, r in enumerate(rows, 1):
        name = r['student'].full_name
        if r.get('is_muqobil') and r.get('muqobil_from'):
            name += f"\n(Muqobil: {r['muqobil_from']})"
        row_data = [
            i,
            name,
            r['total'],
            r['present'],
            r['absent'],
            r['late_count'],
            r['late_mins'],
            r['early_count'],
            r['early_mins'],
            f"{r['percent']}%",
        ]
        if r.get('is_muqobil'):
            row_fill = muqobil_fill
        elif r['percent'] < 70:
            row_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        else:
            row_fill = None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            cell.border = brd
            cell.alignment = center
            if row_fill:
                cell.fill = row_fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f"student_report_{group.pk}_{date_from}_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
