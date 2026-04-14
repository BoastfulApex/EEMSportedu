import json
import calendar
import openpyxl
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from datetime import timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from apps.superadmin.decorators import hr_admin_required, any_admin_required, monitoring_required
from django.db.models import Q, OuterRef, Subquery, Count, F
from django.db.models.functions import TruncMonth, ExtractWeekDay
from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.template import loader
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.generic.edit import DeleteView

from apps.superadmin.models import Administrator, Filial, Weekday
from apps.main.models import Employee, WorkSchedule, Attendance, Schedule, ScheduleDay, ExtraSchedule, SalaryConfig, DailyAttendanceSummary
from apps.main.forms import (
    EmployeeForm, ScheduleForm, AttendanceDateRangeForm, SalaryConfigForm
)


# ============================================================
# HISOBOT YORDAMCHI FUNKSIYALARI
# ============================================================

def _parse_dates(start_date, end_date):
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date,   '%Y-%m-%d').date()
    return start_date, end_date


def _total_minutes(check_in, check_out, date):
    if not check_in or not check_out:
        return 0
    delta = datetime.combine(date, check_out) - datetime.combine(date, check_in)
    return max(0, int(delta.total_seconds() / 60))


GRACE_MINUTES = 15  # 15 daqiqagacha kechikish/erta ketish kechirilaди


def _late_minutes(check_in, schedule_start, date):
    """Kechikish daqiqalari. 15 daqiqagacha grace period."""
    if not check_in:
        return '-'
    diff = int((datetime.combine(date, check_in) -
                datetime.combine(date, schedule_start)).total_seconds() / 60)
    if diff <= 0:
        return 0
    if diff <= GRACE_MINUTES:
        return 0  # Grace period — kechirilaди
    return diff


def _early_leave_minutes(check_out, schedule_end, date):
    """Erta ketish daqiqalari. 15 daqiqagacha grace period."""
    if not check_out:
        return '-'
    diff = int((datetime.combine(date, schedule_end) -
                datetime.combine(date, check_out)).total_seconds() / 60)
    if diff <= 0:
        return 0
    if diff <= GRACE_MINUTES:
        return 0  # Grace period — kechirilaди
    return diff


def _overtime_minutes(check_out, schedule_end, date):
    """Ortiqcha ishlash daqiqalari."""
    if not check_out:
        return 0
    diff = int((datetime.combine(date, check_out) -
                datetime.combine(date, schedule_end)).total_seconds() / 60)
    return max(0, diff)


UZ_DAYS = {
    0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
    3: 'Payshanba', 4: 'Juma', 5: 'Shanba', 6: 'Yakshanba',
}


def _build_daily_report(employees_qs, start_date, end_date):
    """
    Belgilangan davr uchun kunlik davomat yozuvlari.
    Har bir ScheduleDay + Attendance = bitta qator.
    15 daqiqa grace period qo'llaniladi.
    """
    start_date, end_date = _parse_dates(start_date, end_date)
    rows = []
    delta = timedelta(days=1)
    idx = 1

    current = start_date
    while current <= end_date:
        wd_name = calendar.day_name[current.weekday()]
        try:
            weekday_obj = Weekday.objects.get(name_en=wd_name)
        except Weekday.DoesNotExist:
            current += delta
            continue

        for emp in employees_qs:
            if emp.created_at.date() > current:
                continue

            schedule_days = ScheduleDay.objects.filter(
                schedule__employees=emp,
                weekday=weekday_obj
            ).select_related('schedule__location')

            for sd in schedule_days:
                att = Attendance.objects.filter(
                    employee=emp, date=current, location=sd.schedule.location
                ).first()
                if att is None:
                    att = Attendance.objects.filter(
                        employee=emp, date=current
                    ).first()

                check_in  = att.check_in  if att else None
                check_out = att.check_out if att else None
                location  = sd.schedule.location.name if sd.schedule.location else '—'

                worked_min = _total_minutes(check_in, check_out, current)

                late = _late_minutes(check_in, sd.start, current)
                late_min = late if late != '-' else 0

                overtime_min = _overtime_minutes(check_out, sd.end, current)

                rows.append({
                    'index':          idx,
                    'date':           current,
                    'weekday':        UZ_DAYS.get(current.weekday(), wd_name),
                    'employee':       emp,
                    'location':       location,
                    'schedule_start': sd.start,
                    'schedule_end':   sd.end,
                    'status':         'Kelgan' if att else 'Kelmagan',
                    'check_in':       check_in,
                    'check_out':      check_out,
                    'worked_h':       worked_min // 60,
                    'worked_m':       worked_min % 60,
                    'worked_total':   worked_min,
                    'late_min':       late_min,
                    'overtime_min':   overtime_min,
                })
                idx += 1

        current += delta

    return rows


def _build_emp_stats_for_period(employees_qs, start_date, end_date):
    """
    [start_date, end_date] oralig'i uchun har bir xodim bo'yicha xulosa.
    Qaytaradi: list of dict {employee, required_hours, worked_hours,
    progress_pct, late_*, early_*, overtime_*}
    """
    start_date, end_date = _parse_dates(start_date, end_date)
    stats = []
    delta = timedelta(days=1)

    for emp in employees_qs:
        required_minutes = 0
        worked_minutes = 0
        late_total = 0
        early_leave_total = 0
        overtime_total = 0

        current = start_date
        while current <= end_date:
            wd_name = calendar.day_name[current.weekday()]
            try:
                weekday_obj = Weekday.objects.get(name_en=wd_name)
            except Weekday.DoesNotExist:
                current += delta
                continue

            schedule_days = ScheduleDay.objects.filter(
                schedule__employees=emp,
                weekday=weekday_obj
            ).select_related('schedule__location')

            counted_att_ids = set()
            day_late = 0
            day_early = 0
            day_overtime = 0

            for sd in schedule_days:
                # Jadval vaqtiga ko'ra kerakli daqiqalar
                req = int((
                    datetime.combine(current, sd.end) -
                    datetime.combine(current, sd.start)
                ).total_seconds() / 60)
                required_minutes += max(0, req)

                att = Attendance.objects.filter(
                    employee=emp, date=current, location=sd.schedule.location
                ).first()
                if att is None:
                    att = Attendance.objects.filter(
                        employee=emp, date=current
                    ).exclude(id__in=counted_att_ids).first()

                if att and att.id not in counted_att_ids:
                    counted_att_ids.add(att.id)
                    worked_minutes += _total_minutes(att.check_in, att.check_out, current)

                if att:
                    late = _late_minutes(att.check_in, sd.start, current)
                    if late != '-':
                        day_late = max(day_late, late)

                    early = _early_leave_minutes(att.check_out, sd.end, current)
                    if early != '-':
                        day_early = max(day_early, early)

                    day_overtime += _overtime_minutes(att.check_out, sd.end, current)

            late_total += day_late
            early_leave_total += day_early
            overtime_total += day_overtime
            current += delta

        required_hours = required_minutes / 60
        worked_hours = worked_minutes / 60

        if required_hours > 0:
            progress_pct = min(150, round(worked_hours / required_hours * 100, 1))
        else:
            progress_pct = 0

        stats.append({
            'employee':    emp,
            'required_h':  round(required_hours, 1),
            'worked_h':    round(worked_hours, 1),
            'progress_pct': progress_pct,
            'late_h':      late_total // 60,
            'late_m':      late_total % 60,
            'late_total':  late_total,
            'early_h':     early_leave_total // 60,
            'early_m':     early_leave_total % 60,
            'early_total': early_leave_total,
            'overtime_h':  overtime_total // 60,
            'overtime_m':  overtime_total % 60,
            'overtime_total': overtime_total,
        })

    return stats


def _build_day_rows(employee, current, weekday_obj, week_uz):
    """
    Bir xodimning bir kuniga tegishli barcha jadval qatorlarini qaytaradi.
    Yangi Schedule M2M dan foydalanadi.
    """
    rows = []

    # Xodimga biriktirilgan jadvallarning bugungi kunga mos ScheduleDay larini olish
    schedule_days = ScheduleDay.objects.filter(
        schedule__employees=employee,
        weekday=weekday_obj
    ).select_related('schedule__location', 'weekday')

    for sd in schedule_days:
        sch = sd.schedule
        att = Attendance.objects.filter(
            employee=employee, date=current, location=sch.location
        ).first()
        if att is None:
            att = Attendance.objects.filter(employee=employee, date=current).first()

        check_in  = att.check_in  if att else None
        check_out = att.check_out if att else None
        status    = "Kelgan" if att else "Kelmagan"
        worked_min = _total_minutes(check_in, check_out, current)
        worked_str = f"{worked_min // 60}s {worked_min % 60}d" if worked_min else "-"

        rows.append({
            'date':                current,
            'weekday':             week_uz,
            'employee':            employee.name,
            'employee_type':       employee.get_employee_type_display(),
            'schedule_type':       sch.name,
            'location':            sch.location.name if sch.location else '-',
            'schedule_start':      sd.start,
            'schedule_end':        sd.end,
            'status':              status,
            'check_in':            check_in or '-',
            'check_out':           check_out or '-',
            'worked':              worked_str,
            'late_minutes':        _late_minutes(check_in, sd.start, current),
            'early_leave_minutes': _early_leave_minutes(check_out, sd.end, current),
        })

    return rows


# ============================================================
# ASOSIY HISOBOT FUNKSIYALARI
# ============================================================

def build_report(start_date, end_date, filial_id=None):
    start_date, end_date = _parse_dates(start_date, end_date)
    report = []
    delta  = timedelta(days=1)
    current = start_date

    while current <= end_date:
        weekday_name = calendar.day_name[current.weekday()]
        try:
            weekday_obj = Weekday.objects.get(name_en=weekday_name)
            week_uz = weekday_obj.name
        except Weekday.DoesNotExist:
            current += delta
            continue

        # Shu kunda ScheduleDay yozuvi bor xodimlar
        employees_today = Employee.objects.filter(
            schedules__days__weekday=weekday_obj,
            created_at__date__lte=current,
        ).distinct()
        if filial_id:
            employees_today = employees_today.filter(filial_id=int(filial_id))

        for employee in employees_today:
            rows = _build_day_rows(employee, current, weekday_obj, week_uz)
            report.extend(rows)

        current += delta

    for i, row in enumerate(report, 1):
        row['index'] = i
    return report


def build_report_for_employee(employee_id, start_date, end_date):
    start_date, end_date = _parse_dates(start_date, end_date)
    employee = Employee.objects.get(id=employee_id)
    report   = []
    delta    = timedelta(days=1)
    current  = start_date

    while current <= end_date:
        if employee.created_at.date() > current:
            current += delta
            continue

        weekday_name = calendar.day_name[current.weekday()]
        try:
            weekday_obj = Weekday.objects.get(name_en=weekday_name)
            week_uz = weekday_obj.name
        except Weekday.DoesNotExist:
            current += delta
            continue

        rows = _build_day_rows(employee, current, weekday_obj, week_uz)
        report.extend(rows)
        current += delta

    for i, row in enumerate(report, 1):
        row['index'] = i
    return report


# ============================================================
# DASHBOARD
# ============================================================

@any_admin_required
def index(request):
    admin_user = request.admin_user

    # Monitoring admini to'g'ridan-to'g'ri monitoring dashboardga yo'naltiriladi
    if admin_user.is_monitoring and not (admin_user.is_org_admin or admin_user.is_filial_admin
                                          or admin_user.is_hr_admin or admin_user.is_edu_admin):
        return redirect(reverse('monitoring_dashboard'))

    data = {}
    filial = ''
    total_attendance_count = 0
    todays_attendance_count = 0
    early_leave_percent = 0
    late_percent = 0
    late_count = 0
    early_leave_count = 0
    chart_labels = []
    late_values = []
    early_values = []

    tashkent_time = timezone.localtime(timezone.now())

    if admin_user.is_org_admin and 'selected_filial_id' not in request.session:
        request.session['selected_filial_id'] = 'super_admin'

    selected_filial_id = request.session.get('selected_filial_id', 'super_admin')

    if admin_user.is_org_admin:
        filials = Filial.objects.filter(organization=admin_user.organization)
        data['filials'] = filials
        data['selected_filial_id'] = selected_filial_id
        template = 'home/superuser/super_dashboard.html'
        try:
            filial = Filial.objects.get(id=int(selected_filial_id))
        except Exception:
            filial = ''
    else:
        template = 'home/user/staff_dashboard.html'
        filial = admin_user.filial

    if selected_filial_id != 'super_admin' and filial:
        today = timezone.localdate()
        week_start = today - timedelta(days=6)

        todays_attendance_count = Attendance.objects.filter(
            employee__filial=filial, date=today
        ).count()

        attendances = Attendance.objects.filter(
            employee__filial=filial,
            date__range=[week_start, today]
        ).select_related('employee__schedules')

        for att in attendances:
            total_attendance_count += 1
            # Xodimning bugungi ScheduleDay orqali kechikish/erta ketishni hisoblaymiz
            today_wd = att.date.weekday()  # 0=Monday
            emp_day = ScheduleDay.objects.filter(
                schedule__employees=att.employee
            ).select_related('weekday').filter(weekday__name_en__iexact=calendar.day_name[today_wd]).first()
            if emp_day:
                if att.check_in and att.check_in > emp_day.start:
                    late_count += 1
                if att.check_out and att.check_out < emp_day.end:
                    early_leave_count += 1

        if total_attendance_count > 0:
            late_percent = late_count / total_attendance_count * 100
            early_leave_percent = early_leave_count / total_attendance_count * 100

        template = 'home/user/staff_dashboard.html'

    context = {
        'segment': 'dashboard',
        'data': data,
        'filial': filial,
        'tashkent_time': tashkent_time,
        'todays_attendance_count': todays_attendance_count,
        'late_percent': round(late_percent, 1),
        'early_leave_percent': round(early_leave_percent, 1),
        'chart_labels_json': json.dumps(chart_labels),
        'late_values_json': json.dumps(late_values),
        'early_values_json': json.dumps(early_values),
    }
    return HttpResponse(loader.get_template(template).render(context, request))


# ============================================================
# XODIMLAR
# ============================================================

def _get_filial_id(admin_user, request):
    if admin_user.is_org_admin:
        selected = request.session.get('selected_filial_id', 'super_admin')
        if selected == 'super_admin':
            return None
        return int(selected)
    return admin_user.filial.id if admin_user.filial else None


def _base_context(admin_user):
    return {
        'filials': Filial.objects.filter(organization=admin_user.organization)
    }


@hr_admin_required
def employees(request):
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)
    emps = Employee.objects.filter(filial_id=filial_id)
    search_query = request.GET.get('q')
    if search_query:
        emps = emps.filter(Q(name__icontains=search_query))
    paginator = Paginator(emps, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'segment': 'employees',
        'filial': filial.filial_name,
        'tashkent_time': timezone.localtime(timezone.now()),
        'data': data,
    }
    return HttpResponse(loader.get_template('home/user/employees/employees.html').render(context, request))


@hr_admin_required
def employee_create(request):
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)

    if request.method == 'POST':
        emp_form = EmployeeForm(request.POST, request.FILES, filial=filial)
        if emp_form.is_valid():
            employee = emp_form.save(commit=False)
            employee.filial = filial
            employee.save()
            emp_form.save_m2m()
            return redirect('employees')
    else:
        emp_form = EmployeeForm(filial=filial)

    return render(request, 'home/user/employees/employee_create.html', {
        'emp_form': emp_form,
        'filial': filial.filial_name,
        'segment': 'employees',
        'data': data,
        'tashkent_time': timezone.localtime(timezone.now()),
    })


@hr_admin_required
def employee_detail(request, pk):
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    employee = get_object_or_404(Employee, id=pk)
    if employee.filial_id != filial_id:
        return redirect('home')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = employee.filial

    salary_cfg, _ = SalaryConfig.objects.get_or_create(employee=employee)

    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee, filial=filial)
        salary_form = SalaryConfigForm(request.POST, instance=salary_cfg)
        if form.is_valid() and salary_form.is_valid():
            emp = form.save(commit=False)
            if 'image' in request.FILES:
                emp.image = request.FILES['image']
            emp.save()
            form.save_m2m()
            salary_form.save()
            return redirect('employees')
    else:
        form = EmployeeForm(instance=employee, filial=filial)
        salary_form = SalaryConfigForm(instance=salary_cfg)

    return render(request, 'home/user/employees/employee_detail.html', {
        'form': form,
        'salary_form': salary_form,
        'segment': 'employees',
        'employee': employee,
        'filial': employee.filial.filial_name,
        'tashkent_time': timezone.localtime(timezone.now()),
        'data': data,
    })


class EmployeeDelete(DeleteView):
    model = Employee
    fields = '__all__'
    success_url = reverse_lazy('employees')


# ============================================================
# JADVAL SHABLONLARI (yangi)
# ============================================================

@hr_admin_required
def schedules(request):
    """Jadval shablonlari ro'yxati"""
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)
    search_query = request.GET.get('q', '')

    qs = Schedule.objects.filter(filial_id=filial_id).prefetch_related('days__weekday', 'location').order_by('name')
    if search_query:
        qs = qs.filter(name__icontains=search_query)

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'home/user/workschedule/schedules.html', {
        'page_obj': page_obj,
        'segment': 'schedules',
        'filial': filial.filial_name,
        'tashkent_time': timezone.localtime(timezone.now()),
        'data': data,
        'search_query': search_query,
    })


def _save_schedule_days(schedule, post_data):
    """POST dan kun vaqtlarini o'qib ScheduleDay larni yaratadi/yangilaydi."""
    all_weekdays = Weekday.objects.all()
    for wd in all_weekdays:
        key = f'day_{wd.id}'
        start_val = post_data.get(f'{key}_start', '').strip()
        end_val   = post_data.get(f'{key}_end', '').strip()
        if start_val and end_val:
            ScheduleDay.objects.update_or_create(
                schedule=schedule, weekday=wd,
                defaults={'start': start_val, 'end': end_val}
            )
        else:
            ScheduleDay.objects.filter(schedule=schedule, weekday=wd).delete()


@hr_admin_required
def schedule_create(request):
    """Yangi jadval shabloni yaratish"""
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)
    weekdays = Weekday.objects.all()

    if request.method == 'POST':
        form = ScheduleForm(request.POST, filial=filial)
        if form.is_valid():
            sch = form.save(commit=False)
            sch.filial = filial
            sch.save()
            _save_schedule_days(sch, request.POST)
            return redirect('schedules')
    else:
        form = ScheduleForm(filial=filial)

    return render(request, 'home/user/workschedule/schedule_create.html', {
        'form': form,
        'weekdays': weekdays,
        'filial': filial.filial_name,
        'segment': 'schedules',
        'tashkent_time': timezone.localtime(timezone.now()),
        'data': data,
    })


@hr_admin_required
def schedule_detail(request, pk):
    """Jadval shablonini tahrirlash"""
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    schedule = get_object_or_404(Schedule, id=pk, filial_id=filial_id)
    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)
    weekdays = Weekday.objects.all()

    # Mavjud kunlar dict: {weekday_id: ScheduleDay}
    existing_days = {sd.weekday_id: sd for sd in schedule.days.select_related('weekday')}

    if request.method == 'POST':
        form = ScheduleForm(request.POST, instance=schedule, filial=filial)
        if form.is_valid():
            form.save()
            _save_schedule_days(schedule, request.POST)
            return redirect('schedules')
    else:
        form = ScheduleForm(instance=schedule, filial=filial)

    return render(request, 'home/user/workschedule/schedule_detail.html', {
        'form': form,
        'schedule': schedule,
        'weekdays': weekdays,
        'existing_days': existing_days,
        'filial': filial.filial_name,
        'segment': 'schedules',
        'tashkent_time': timezone.localtime(timezone.now()),
        'data': data,
    })


class ScheduleDelete(DeleteView):
    model = Schedule
    success_url = reverse_lazy('schedules')
    template_name = 'main/schedule_confirm_delete.html'


# ============================================================
# HISOBOT
# ============================================================

@any_admin_required
def get_report_date(request):
    """
    Yagona hisobot sahifasi.
    ?type=monthly   → joriy oy boshi – bugun
    ?type=weekly    → joriy hafta dushanba – bugun
    ?type=date_range → foydalanuvchi sanani tanlaydi
    Har bir holat uchun xodimlar bo'yicha jadval ko'rinishida xulosa.
    """
    admin_user = request.admin_user
    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    data = {'filials': _base_context(admin_user)['filials']}
    filial = Filial.objects.get(id=filial_id)
    today = timezone.localdate()

    report_type = request.GET.get('type', 'monthly')

    UZ_MONTHS = {
        1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
        5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
        9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr',
    }

    if report_type == 'monthly':
        start_date = today.replace(day=1)
        end_date = today
        period_label = f"{UZ_MONTHS[today.month]} {today.year} ({start_date} – {end_date})"

    elif report_type == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
        period_label = f"{start_date} – {end_date} (joriy hafta)"

    else:  # date_range yoki daily — foydalanuvchi sanani tanlaydi
        start_str = request.GET.get('start_date', '')
        end_str   = request.GET.get('end_date', '')
        if start_str and end_str:
            try:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date   = datetime.strptime(end_str,   '%Y-%m-%d').date()
            except ValueError:
                start_date = today
                end_date   = today
        else:
            start_date = today
            end_date   = today
        period_label = f"{start_date} – {end_date}"

    employees_qs = Employee.objects.filter(filial_id=filial_id).order_by('name')

    if report_type == 'daily':
        daily_rows = _build_daily_report(employees_qs, start_date, end_date)
        emp_stats  = []
    else:
        emp_stats  = _build_emp_stats_for_period(employees_qs, start_date, end_date)
        daily_rows = []

    return render(request, 'home/user/report/get_report_date.html', {
        'emp_stats':    emp_stats,
        'daily_rows':   daily_rows,
        'report_type':  report_type,
        'period_label': period_label,
        'start_date':   start_date,
        'end_date':     end_date,
        'segment':      'report',
        'tashkent_time': timezone.localtime(timezone.now()),
        'filial':       filial.filial_name,
        'data':         data,
        'today':        today,
    })


@any_admin_required
def download_excel(request):
    admin_user = request.admin_user

    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if not start_date or not end_date:
        return redirect('home_get_dates')

    report_data = build_report(start_date=start_date, end_date=end_date, filial_id=filial_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat hisoboti"
    headers = ['T/r', 'Sana', 'Hafta kuni', 'Xodim', 'Turi', 'Holati',
               'Jadval', 'Lokatsiya', 'Jadval (boshlanish)', 'Jadval (tugash)',
               'Kirish', 'Chiqish', 'Jami ish vaqti', 'Kechikdi (daqiqa)', 'Erta ketdi (daqiqa)']
    ws.append(headers)
    for row in report_data:
        ws.append([
            row['index'], str(row['date']), row['weekday'],
            row['employee'], row['employee_type'], row['status'],
            row['schedule_type'], row['location'],
            str(row['schedule_start']), str(row['schedule_end']),
            str(row['check_in']), str(row['check_out']),
            row['worked'], row['late_minutes'], row['early_leave_minutes'],
        ])
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=hisobot.xlsx'
    wb.save(response)
    return response


@any_admin_required
def report_download_excel(request):
    """
    Hisobot turига ko'ra Excel yuklab olish.
    ?type=monthly|weekly|date_range  → xodimlar bo'yicha xulosa
    ?type=daily                      → kunlik davomat jadvali
    """
    admin_user = request.admin_user
    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    today = timezone.localdate()
    report_type = request.GET.get('type', 'monthly')
    start_str   = request.GET.get('start_date', '')
    end_str     = request.GET.get('end_date', '')

    if report_type == 'monthly':
        start_date = today.replace(day=1)
        end_date   = today
    elif report_type == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date   = today
    else:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date   = datetime.strptime(end_str,   '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return redirect('home_get_dates')

    employees_qs = Employee.objects.filter(filial_id=filial_id).order_by('name')
    wb = openpyxl.Workbook()
    ws = wb.active

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    center = Alignment(horizontal='center', vertical='center')

    if report_type == 'daily':
        ws.title = "Kunlik davomat"
        headers = [
            '#', 'Sana', 'Hafta kuni', 'Xodim', 'Turi',
            'Lokatsiya', 'Jadval boshi', 'Jadval oxiri',
            'Holat', 'Kirish', 'Chiqish',
            'Ishlagan (soat)', 'Kechikish (daqiqa)', 'Ortiqcha (daqiqa)'
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        rows = _build_daily_report(employees_qs, start_date, end_date)
        for r in rows:
            ws.append([
                r['index'],
                str(r['date']),
                r['weekday'],
                r['employee'].name,
                r['employee'].get_employee_type_display(),
                r['location'],
                str(r['schedule_start']),
                str(r['schedule_end']),
                r['status'],
                str(r['check_in']) if r['check_in'] else '—',
                str(r['check_out']) if r['check_out'] else '—',
                round((r['worked_h'] * 60 + r['worked_m']) / 60, 2),
                r['late_min'],
                r['overtime_min'],
            ])
    else:
        ws.title = "Xodimlar xulosasi"
        type_labels = {
            'monthly': 'Oylik', 'weekly': 'Haftalik', 'date_range': 'Sana bo\'yicha'
        }
        headers = [
            '#', 'Xodim', 'Turi',
            'Kerakli soat', 'Ishlagan soat', 'Bajarilish %',
            'Kechikish (daqiqa)', 'Erta ketish (daqiqa)', 'Ortiqcha (daqiqa)'
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        stats = _build_emp_stats_for_period(employees_qs, start_date, end_date)
        for i, e in enumerate(stats, 1):
            ws.append([
                i,
                e['employee'].name,
                e['employee'].get_employee_type_display(),
                e['required_h'],
                e['worked_h'],
                e['progress_pct'],
                e['late_total'],
                e['early_total'],
                e['overtime_total'],
            ])

    # Ustun kengligini avtomatik moslashtirish
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    filename = f"hisobot_{report_type}_{start_date}_{end_date}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@any_admin_required
def employee_report(request, pk):
    """
    Xodimning shaxsiy hisobot sahifasi.
    Doim ko'rsatiladi: Haftalik va Oylik xulosa cartlari.
    Ixtiyoriy: ?type=date_range | daily — sana tanlash + jadval.
    """
    admin_user = request.admin_user
    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    employee = get_object_or_404(Employee, id=pk)
    data = {'filials': _base_context(admin_user)['filials']}
    today = timezone.localdate()

    # ── Haftalik va oylik avtomatik cardlar ──
    week_start  = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    emp_qs = Employee.objects.filter(id=pk)

    weekly_list  = _build_emp_stats_for_period(emp_qs, week_start, today)
    monthly_list = _build_emp_stats_for_period(emp_qs, month_start, today)

    weekly_stats  = weekly_list[0]  if weekly_list  else None
    monthly_stats = monthly_list[0] if monthly_list else None

    UZ_MONTHS = {
        1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
        5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
        9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr',
    }

    # ── Qo'shimcha hisobot (ixtiyoriy) ──
    report_type = request.GET.get('type', '')
    start_date = end_date = period_label = None
    emp_stats = daily_rows = []

    if report_type in ('date_range', 'daily'):
        start_str = request.GET.get('start_date', '')
        end_str   = request.GET.get('end_date', '')
        if start_str and end_str:
            try:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date   = datetime.strptime(end_str,   '%Y-%m-%d').date()
            except ValueError:
                start_date = end_date = today
        else:
            start_date = end_date = today
        period_label = f"{start_date} – {end_date}"

        if report_type == 'daily':
            daily_rows = _build_daily_report(emp_qs, start_date, end_date)
        else:
            emp_stats = _build_emp_stats_for_period(emp_qs, start_date, end_date)

    return render(request, 'home/user/report/employee_report.html', {
        'employee':      employee,
        'weekly_stats':  weekly_stats,
        'monthly_stats': monthly_stats,
        'week_start':    week_start,
        'month_start':   month_start,
        'month_label':   f"{UZ_MONTHS[today.month]} {today.year}",
        'emp_stats':     emp_stats,
        'daily_rows':    daily_rows,
        'report_type':   report_type,
        'period_label':  period_label,
        'start_date':    start_date or today,
        'end_date':      end_date or today,
        'today':         today,
        'segment':       'employees',
        'data':          data,
        'tashkent_time': timezone.localtime(timezone.now()),
        'filial':        employee.filial.filial_name if employee.filial else '',
    })


@any_admin_required
def employee_download_excel(request, pk):
    """
    Xodimning to'liq hisobotini bitta report.xlsx ga yozadi.
    Varaqlar: Oylik, Haftalik, va (ixtiyoriy) Kunlik yoki Sana bo'yicha.
    """
    employee = get_object_or_404(Employee, id=pk)
    today    = timezone.localdate()
    emp_qs   = Employee.objects.filter(id=pk)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # default sheet ni o'chirish

    def _style_header(ws, headers):
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 38)

    def _write_summary_sheet(ws, stats_list, period_str):
        """Xulosa varag'i: bir qator xodim uchun."""
        _style_header(ws, [
            'Davr', 'Xodim', 'Turi',
            'Kerakli (soat)', 'Ishlagan (soat)', 'Bajarilish (%)',
            'Kechikish (daqiqa)', 'Erta ketish (daqiqa)', 'Ortiqcha (daqiqa)',
        ])
        for e in stats_list:
            ws.append([
                period_str,
                e['employee'].name,
                e['employee'].get_employee_type_display(),
                e['required_h'],
                e['worked_h'],
                e['progress_pct'],
                e['late_total'],
                e['early_total'],
                e['overtime_total'],
            ])
        _autofit(ws)

    def _write_daily_sheet(ws, rows):
        """Kunlik davomat varag'i."""
        _style_header(ws, [
            '#', 'Sana', 'Hafta kuni', 'Xodim', 'Turi',
            'Lokatsiya', 'Jadval boshi', 'Jadval oxiri',
            'Holat', 'Kirish', 'Chiqish',
            'Ishlagan (soat)', 'Kechikish (daqiqa)', 'Ortiqcha (daqiqa)',
        ])
        for r in rows:
            ws.append([
                r['index'],
                str(r['date']),
                r['weekday'],
                r['employee'].name,
                r['employee'].get_employee_type_display(),
                r['location'],
                str(r['schedule_start']),
                str(r['schedule_end']),
                r['status'],
                str(r['check_in'])  if r['check_in']  else '—',
                str(r['check_out']) if r['check_out'] else '—',
                round((r['worked_h'] * 60 + r['worked_m']) / 60, 2),
                r['late_min'],
                r['overtime_min'],
            ])
        _autofit(ws)

    # ── 1. Oylik varaq ──
    month_start = today.replace(day=1)
    ws_monthly  = wb.create_sheet("Oylik")
    monthly     = _build_emp_stats_for_period(emp_qs, month_start, today)
    _write_summary_sheet(ws_monthly, monthly, f"{month_start} – {today}")

    # ── 2. Haftalik varaq ──
    week_start = today - timedelta(days=today.weekday())
    ws_weekly  = wb.create_sheet("Haftalik")
    weekly     = _build_emp_stats_for_period(emp_qs, week_start, today)
    _write_summary_sheet(ws_weekly, weekly, f"{week_start} – {today}")

    # ── 3. Qo'shimcha varaq (ixtiyoriy) ──
    report_type = request.GET.get('type', '')
    start_str   = request.GET.get('start_date', '')
    end_str     = request.GET.get('end_date', '')

    if report_type in ('date_range', 'daily') and start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date   = datetime.strptime(end_str,   '%Y-%m-%d').date()
            period_str = f"{start_date} – {end_date}"

            if report_type == 'daily':
                ws_extra = wb.create_sheet("Kunlik")
                rows = _build_daily_report(emp_qs, start_date, end_date)
                _write_daily_sheet(ws_extra, rows)
            else:
                ws_extra = wb.create_sheet("Sana bo'yicha")
                stats = _build_emp_stats_for_period(emp_qs, start_date, end_date)
                _write_summary_sheet(ws_extra, stats, period_str)
        except ValueError:
            pass

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb.save(response)
    return response


# ============================================================
# OYLIK XULOSA HISOBOTI
# ============================================================

@any_admin_required
def monthly_summary(request):
    """Eski URL — asosiy hisobot sahifasiga yo'naltiradi."""
    return redirect(reverse('home_get_dates') + '?type=monthly')


# ============================================================
# MAOSH BOSHQARUVI
# ============================================================

@hr_admin_required
def salary_list(request):
    """Barcha xodimlarning oylik maosh sozlamalarini ko'rsatish."""
    admin_user = request.admin_user
    filial_id = _get_filial_id(admin_user, request)
    if filial_id is None:
        return redirect('/home/')

    filial = Filial.objects.get(id=filial_id)
    employees_qs = Employee.objects.filter(filial_id=filial_id).order_by('name')

    salary_configs = {
        sc.employee_id: sc
        for sc in SalaryConfig.objects.filter(employee__in=employees_qs)
    }

    emp_data = []
    for emp in employees_qs:
        cfg = salary_configs.get(emp.id)
        emp_data.append({
            'employee'      : emp,
            'monthly_hours'  : cfg.monthly_hours   if cfg else None,
            'monthly_salary' : cfg.monthly_salary  if cfg else None,
            'hourly_rate'    : cfg.hourly_rate      if cfg else None,
        })

    return render(request, 'home/user/salary/salary_list.html', {
        'emp_data' : emp_data,
        'filial'   : filial.filial_name,
        'segment'  : 'salary',
        'data'     : {'filials': _base_context(admin_user)['filials']},
    })


@hr_admin_required
@require_POST
def salary_update(request, pk):
    """AJAX: xodim maosh konfiguratsiyasini saqlash."""
    admin_user = request.admin_user
    filial_id = _get_filial_id(admin_user, request)

    employee = get_object_or_404(Employee, id=pk)
    if employee.filial_id != filial_id:
        return JsonResponse({'ok': False, 'error': "Ruxsat yo'q"}, status=403)

    salary_cfg, _ = SalaryConfig.objects.get_or_create(employee=employee)
    form = SalaryConfigForm(request.POST, instance=salary_cfg)
    if form.is_valid():
        cfg = form.save()
        return JsonResponse({
            'ok'            : True,
            'monthly_hours'  : float(cfg.monthly_hours),
            'monthly_salary' : float(cfg.monthly_salary),
            'hourly_rate'    : round(cfg.hourly_rate, 0),
        })
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


# ============================================================
# MONITORING BO'LIMI — TINGLOVCHILAR HISOBOTLARI
# ============================================================

def _get_attendance_limit(admin_user, filial_id):
    """Filial yoki tashkilot uchun davomat limitini qaytaradi."""
    from apps.students.models import AttendanceLimit
    limit = None
    if filial_id:
        limit = AttendanceLimit.objects.filter(
            organization=admin_user.organization, filial_id=filial_id
        ).first()
    if not limit:
        limit = AttendanceLimit.objects.filter(
            organization=admin_user.organization, filial=None
        ).first()
    return limit


def _compute_student_stats(student, group, para_hours):
    """
    Tinglovchining davomat statistikasini hisoblaydi.

    Qoidalar:
    - Faqat BUGUNGI KUNA QADAR bo'lgan va smena belgilangan dars kunlari hisobga olinadi.
    - Har bir para alohida tekshiriladi:
        * check_in > para_start + 40 daq → o'sha paraga kelmagan (missed)
        * check_in <= para_start + 40 daq → o'sha paraga kelgan
    - Barcha paralarga 40+ daqiqa kech qolgan kun → kelmadi (absent) hisoblanadi.
    - Davomat yozilmagan dars kunlari → kelmadi + barcha paralari missed.
    """
    from apps.students.models import StudentAttendance, GroupLesson
    from django.utils import timezone
    import datetime as _dt

    today = timezone.localdate()
    LATE_THRESHOLD = 40  # daqiqa

    lessons_qs = GroupLesson.objects.filter(
        group=group, smena__isnull=False, date__lte=today
    ).select_related('smena')
    lesson_map = {lesson.date: lesson for lesson in lessons_qs}
    scheduled_dates = set(lesson_map.keys())

    att_qs = StudentAttendance.objects.filter(student=student, group=group)
    att_by_date = {att.date: att for att in att_qs}

    present      = 0
    late         = 0
    absent       = 0
    excused      = 0
    missed_paras = 0

    for date, lesson in lesson_map.items():
        smena = lesson.smena
        para_starts = [smena.para1_start]
        if smena.para2_start:
            para_starts.append(smena.para2_start)
        if smena.para3_start:
            para_starts.append(smena.para3_start)
        para_count = len(para_starts)

        att = att_by_date.get(date)

        if att is None:
            absent       += 1
            missed_paras += para_count
            continue

        if att.status == 'excused':
            excused      += 1
            missed_paras += para_count
            continue

        if not att.check_in or att.status == 'absent':
            absent       += 1
            missed_paras += para_count
            continue

        # Har bir parani alohida tekshir
        check_in_dt       = _dt.datetime.combine(date, att.check_in)
        has_checkout      = bool(att.check_out)
        found_kelgan_para = False
        paras_present_today = 0
        paras_missed_today  = 0

        for p in para_starts:
            para_dt = _dt.datetime.combine(date, p)
            if check_in_dt > para_dt + _dt.timedelta(minutes=LATE_THRESHOLD):
                paras_missed_today += 1
            elif not has_checkout and found_kelgan_para:
                # Check_out yo'q + kelgan para topilgan → keyingisi absent
                paras_missed_today += 1
            else:
                paras_present_today += 1
                found_kelgan_para    = True

        if paras_present_today == 0:
            # Barcha paralarga kech yoki check_out yo'q → kelmadi
            absent       += 1
            missed_paras += para_count
        else:
            present      += 1
            missed_paras += paras_missed_today
            if att.late_minutes and att.late_minutes > 0:
                late += 1

    total        = len(scheduled_dates)
    missed_hours = round(missed_paras * para_hours, 1)
    pct          = round(present / total * 100) if total > 0 else 0

    return {
        'total':        total,
        'present':      present,
        'late':         late,
        'absent':       absent,
        'excused':      excused,
        'missed_paras': missed_paras,
        'missed_hours': missed_hours,
        'pct':          pct,
    }


def _build_exceeded_students(groups_qs, limit):
    """Limitdan oshgan tinglovchilar ro'yxatini qaytaradi."""
    if not limit:
        return []

    exceeded = []
    seen = set()

    for group in groups_qs.prefetch_related('students', 'direction'):
        for student in group.students.all():
            key = (student.id, group.id)
            if key in seen:
                continue
            seen.add(key)

            stats = _compute_student_stats(student, group, limit.para_hours)

            if stats['missed_hours'] > limit.max_missed_hours:
                exceeded.append({
                    'student':      student,
                    'group':        group,
                    'direction':    group.direction,
                    'missed_paras': stats['missed_paras'],
                    'missed_hours': stats['missed_hours'],
                    'max_hours':    limit.max_missed_hours,
                    'over_hours':   round(stats['missed_hours'] - limit.max_missed_hours, 1),
                    'pct':          stats['pct'],
                })

    exceeded.sort(key=lambda r: r['missed_hours'], reverse=True)
    return exceeded


@monitoring_required
def monitoring_dashboard(request):
    """Monitoring dashboard: guruhlar, tinglovchilar, kechikish foizi."""
    from apps.students.models import Group, StudentAttendance, Direction
    from django.db.models import Count, Q

    admin_user = request.admin_user
    filial_id  = _get_filial_id(admin_user, request)

    groups_qs = Group.objects.filter(organization=admin_user.organization)
    if filial_id:
        groups_qs = groups_qs.filter(filial_id=filial_id)

    groups_count = groups_qs.count()

    # Noyob tinglovchilar soni
    student_ids = set()
    for g in groups_qs.prefetch_related('students'):
        student_ids.update(g.students.values_list('id', flat=True))
    students_count = len(student_ids)

    # Kechikish foizi (barcha vaqt uchun)
    att_qs    = StudentAttendance.objects.filter(group__in=groups_qs)
    total_att = att_qs.count()
    late_att  = att_qs.filter(Q(late_minutes__gt=0) | Q(status='late')).count()
    late_pct  = round(late_att / total_att * 100, 1) if total_att > 0 else 0

    # Limitdan oshgan tinglovchilar + jami qoldirilgan soat
    limit          = _get_attendance_limit(admin_user, filial_id)
    para_hours     = limit.para_hours if limit else 2.0
    exceeded_list  = _build_exceeded_students(groups_qs, limit)
    exceeded_count = len(exceeded_list)

    # Jami qoldirilgan soat (barcha guruh tinglovchilari bo'yicha)
    total_missed_hours = 0.0
    seen_keys = set()
    for group in groups_qs.prefetch_related('students'):
        for student in group.students.all():
            key = (student.id, group.id)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            stats = _compute_student_stats(student, group, para_hours)
            total_missed_hours += stats['missed_hours']
    total_missed_hours = round(total_missed_hours, 1)

    # So'nggi 30 kunlik trend (kunlik davomat soni)
    from datetime import date, timedelta
    today      = date.today()
    trend_data = []
    for i in range(29, -1, -1):
        d      = today - timedelta(days=i)
        count  = att_qs.filter(date=d, status__in=['present', 'late']).count()
        trend_data.append({'date': str(d), 'count': count})

    return render(request, 'monitoring/dashboard.html', {
        'segment':        'monitoring_dashboard',
        'groups_count':   groups_count,
        'students_count': students_count,
        'late_pct':       late_pct,
        'total_att':      total_att,
        'trend_data':     json.dumps(trend_data),
        'exceeded_count':      exceeded_count,
        'limit':               limit,
        'total_missed_hours':  total_missed_hours,
        'data':                {'filials': _base_context(admin_user)['filials']},
        'tashkent_time':  timezone.localtime(timezone.now()),
    })


@monitoring_required
def monitoring_reports(request):
    """Tinglovchilar davomati hisoboti — yo'nalish, guruh, yil, oy filtrlari."""
    from apps.students.models import Group, Direction, MONTH_CHOICES

    admin_user = request.admin_user
    filial_id  = _get_filial_id(admin_user, request)

    # Filter uchun ma'lumotlar
    base_filter = {'organization': admin_user.organization}
    if filial_id:
        base_filter['filial_id'] = filial_id

    directions = Direction.objects.filter(**base_filter).order_by('name')
    all_groups = Group.objects.filter(**base_filter).order_by('name')
    years      = sorted(all_groups.values_list('year', flat=True).distinct())

    # GET parametrlar
    direction_id = request.GET.get('direction', '')
    group_id     = request.GET.get('group', '')
    year         = request.GET.get('year', '')
    month        = request.GET.get('month', '')

    # Guruhlarni filtrlash
    filtered_groups = all_groups
    if direction_id:
        filtered_groups = filtered_groups.filter(direction_id=direction_id)
    if year:
        filtered_groups = filtered_groups.filter(year=year)
    if month:
        filtered_groups = filtered_groups.filter(month=month)
    if group_id:
        filtered_groups = filtered_groups.filter(id=group_id)

    # Hisobot qatorlari
    report_rows = []
    any_filter  = any([direction_id, group_id, year, month])

    limit = _get_attendance_limit(admin_user, filial_id)
    para_hours = limit.para_hours if limit else 2.0
    max_hours  = limit.max_missed_hours if limit else None

    if any_filter:
        for group in filtered_groups.prefetch_related('students', 'direction'):
            for student in group.students.all():
                s = _compute_student_stats(student, group, para_hours)
                exceeded = max_hours is not None and s['missed_hours'] > max_hours
                report_rows.append({
                    'student':      student,
                    'group':        group,
                    'direction':    group.direction,
                    'year':         group.year,
                    'month':        group.get_month_display(),
                    'total':        s['total'],
                    'present':      s['present'],
                    'late':         s['late'],
                    'absent':       s['absent'],
                    'excused':      s['excused'],
                    'missed_paras': s['missed_paras'],
                    'missed_hours': s['missed_hours'],
                    'pct':          s['pct'],
                    'exceeded':     exceeded,
                })
        report_rows.sort(key=lambda r: r['pct'])

    return render(request, 'monitoring/reports.html', {
        'segment':       'monitoring_reports',
        'directions':    directions,
        'all_groups':    all_groups,
        'years':         years,
        'months':        MONTH_CHOICES,
        'report_rows':   report_rows,
        'any_filter':    any_filter,
        'f_direction':   direction_id,
        'f_group':       group_id,
        'f_year':        year,
        'f_month':       month,
        'limit':         limit,
        'para_hours':    para_hours,
        'max_hours':     max_hours,
        'data':          {'filials': _base_context(admin_user)['filials']},
        'tashkent_time': timezone.localtime(timezone.now()),
    })


@monitoring_required
def monitoring_limit_settings(request):
    """Davomat limiti sozlamalari."""
    from apps.students.models import AttendanceLimit

    admin_user = request.admin_user
    filial_id  = _get_filial_id(admin_user, request)

    limit, _ = AttendanceLimit.objects.get_or_create(
        organization=admin_user.organization,
        filial_id=filial_id if filial_id else None,
    )

    if request.method == 'POST':
        try:
            para_hours     = float(request.POST.get('para_hours', 2.0))
            max_missed_hours = float(request.POST.get('max_missed_hours', 20.0))
            if para_hours <= 0 or max_missed_hours <= 0:
                raise ValueError
            limit.para_hours       = para_hours
            limit.max_missed_hours = max_missed_hours
            limit.save()
            return redirect(reverse('monitoring_limit_settings') + '?saved=1')
        except (ValueError, TypeError):
            pass

    saved = request.GET.get('saved') == '1'
    return render(request, 'monitoring/limit_settings.html', {
        'segment':   'monitoring_limit',
        'limit':     limit,
        'saved':     saved,
        'data':      {'filials': _base_context(admin_user)['filials']},
        'tashkent_time': timezone.localtime(timezone.now()),
    })


@monitoring_required
def monitoring_exceeded(request):
    """Limitdan oshgan tinglovchilar ro'yxati."""
    from apps.students.models import Group

    admin_user = request.admin_user
    filial_id  = _get_filial_id(admin_user, request)

    groups_qs = Group.objects.filter(organization=admin_user.organization)
    if filial_id:
        groups_qs = groups_qs.filter(filial_id=filial_id)

    limit         = _get_attendance_limit(admin_user, filial_id)
    exceeded_list = _build_exceeded_students(groups_qs, limit)

    return render(request, 'monitoring/exceeded.html', {
        'segment':        'monitoring_exceeded',
        'exceeded_list':  exceeded_list,
        'limit':          limit,
        'data':           {'filials': _base_context(admin_user)['filials']},
        'tashkent_time':  timezone.localtime(timezone.now()),
    })

